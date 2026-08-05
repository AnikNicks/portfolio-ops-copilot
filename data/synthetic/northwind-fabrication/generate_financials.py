"""One-off generator for the intentionally messy financials_raw.xlsx synthetic fixture for the
second data room (northwind-fabrication). Deliberately different messiness than
acme-distribution's fixture: thousands-separator revenue strings and a GBP (not USD) currency
flag, instead of formula-as-text COGS cells - proves pipeline/normalize.py generalizes rather
than being hand-tuned to one export's quirks.

Run once: python generate_financials.py
Not part of the pipeline itself - pipeline/normalize.py consumes the output, not this script.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUT = Path(__file__).parent / "financials_raw.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "P&L FY25"

ws.merge_cells("A1:H1")
ws["A1"] = "Northwind Fabrication Ltd. - Monthly P&L (unaudited, CAD unless noted)"
ws["A1"].font = Font(bold=True, size=13)

headers = ["Month", "Revenue", "COGS", "Gross Profit", "OpEx", "EBITDA", "AR (days)", "Notes"]
ws.append([])
ws.append(headers)
for c in ws[3]:
    c.font = Font(bold=True)

# 12 months with deliberate messiness distinct from acme-distribution's fixture:
# - two months' Revenue is a thousands-separated string ("894,000") instead of a plain number
# - two months are reported in GBP (a UK export order), not CAD, flagged only via Notes
# - one blank row in the middle
# - COGS creeps up relative to revenue across the year (the margin-leakage signal), and AR days
#   climb steadily (the working-capital signal)
rows = [
    ["Jan-25", 845000, 590000, 255000, 155000, 100000, 38, ""],
    ["Feb-25", 812000, 570000, 242000, 152000, 90000, 40, ""],
    ["Mar-25", 858000, 605000, 253000, 158000, 95000, 42, ""],
    ["Apr-25", 865000, 615000, 250000, 160000, 90000, 44, ""],
    [None, None, None, None, None, None, None, None],
    ["May-25", 871000, 622000, 249000, 161000, 88000, 45, ""],
    ["Jun-25", "894,000", 645000, 249000, 163000, 86000, 47, "GBP - UK export order, not converted"],
    ["Jul-25", "901,000", 662000, 239000, 166000, 73000, 49, "GBP - UK export order, not converted"],
    ["Aug-25", 887000, 655000, 232000, 168000, 64000, 52, "COGS ratio elevated - new subcontractor"],
    ["Sep-25", 902000, 671000, 231000, 171000, 60000, 55, ""],
    ["Oct-25", 915000, 688000, 227000, 174000, 53000, 58, ""],
    ["Nov-25", 928000, 702000, 226000, 177000, 49000, 61, ""],
    ["Dec-25", 940000, 718000, 222000, 180000, 42000, 65, "AR aging up sharply"],
]
for r in rows:
    ws.append(r)

for col, width in zip("ABCDEFGH", [10, 12, 12, 13, 12, 11, 10, 30], strict=True):
    ws.column_dimensions[col].width = width

wb.save(OUT)
print(f"wrote {OUT}")
