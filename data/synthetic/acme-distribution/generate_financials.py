"""One-off generator for the intentionally messy financials_raw.xlsx synthetic fixture.

Run once: python generate_financials.py
Not part of the pipeline itself - pipeline/normalize.py consumes the output, not this script.
"""

from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path

OUT = Path(__file__).parent / "financials_raw.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "P&L FY25-26"

# Title row, merged - a common real-world export quirk that breaks naive header parsing
ws.merge_cells("A1:H1")
ws["A1"] = "Acme Distribution Inc. - Monthly P&L (unaudited, CAD unless noted)"
ws["A1"].font = Font(bold=True, size=13)

# Inconsistent header row: mixed casing, abbreviations, a stray merged cell
headers = ["Month", "Revenue", "COGS", "Gross Profit", "OpEx", "EBITDA", "AR (days)", "Notes"]
ws.append([])
ws.append(headers)
for c in ws[3]:
    c.font = Font(bold=True)

# 12 months of data with deliberate messiness:
# - Q3 (Dec-Feb) revenue is in USD, not CAD, unlabeled except in the "Notes" column
# - two rows have COGS as a formula-looking string instead of a number (export artifact)
# - one blank row in the middle (common in Excel exports)
# - Gross Profit column sometimes pre-computed, sometimes not, and disagrees with Revenue-COGS in two rows (rounding/entry error - this IS the margin leakage signal)
rows = [
    ["Mar-25", 612000, 398000, 214000, 165000, 49000, 41, ""],
    ["Apr-25", 598000, 389000, 209000, 168000, 41000, 44, ""],
    ["May-25", 634000, 412000, 222000, 171000, 51000, 39, ""],
    ["Jun-25", 655000, 427000, 228000, 174000, 54000, 42, ""],
    ["Jul-25", 641000, 417000, 224000, 176000, 48000, 47, ""],
    ["Aug-25", 609000, 396000, 213000, 179000, 34000, 51, "AR aging up"],
    [None, None, None, None, None, None, None, None],
    ["Sep-25", 627000, 408000, 219000, 182000, 37000, 53, "AR aging up"],
    ["Oct-25", 618000, 402000, 216000, 184000, 32000, 58, "AR aging up"],
    ["Nov-25", 601000, "601000*0.651", 210000, 187000, 23000, 61, "COGS % creeping"],
    ["Dec-25", 489000, "489000*0.668", 162000, 189000, -27000, 66, "USD - not converted"],
    ["Jan-26", 502000, 335000, 167000, 191000, -24000, 69, "USD - not converted"],
    ["Feb-26", 511000, 341000, 170000, 193000, -23000, 71, "USD - not converted"],
]
for r in rows:
    ws.append(r)

for col, width in zip("ABCDEFGH", [10, 12, 12, 13, 12, 11, 10, 22]):
    ws.column_dimensions[col].width = width

wb.save(OUT)
print(f"wrote {OUT}")
